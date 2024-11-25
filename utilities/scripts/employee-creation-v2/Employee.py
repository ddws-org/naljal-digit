
#!/usr/bin/python
# -*- coding: utf-8 -*-
import csv
from copy import deepcopy

import requests
import openpyxl
import time
import datetime
import subprocess
import os

host = 'https://naljalseva.jjm.gov.in/assam/'
username = ''  # superuser mobile no
password = ''  # superuser password
stateTenantId='as'
path='updated-user.xlsx'
#cmd=['kubectl', '-it', 'exec', 'kafka-v2-0',  '-n', 'mgramseva', '--', 'kafka-console-producer', '--broker-list', 'localhost:9092', '--topic',  'egov.core.notification.sms']
cmd = "kubectl -it exec kafka-v2-0  -n mgramseva -- kafka-console-producer --broker-list localhost:9092 --topic  egov.core.notification.sms <sms.txt"
iphindi='{"mobileNumber":"{PHNO}","message":"प्रिय {USER}, आपको पंजाब के mGramSeva एप्लिकेशन में आमंत्रित किया गया है। कृपया {LINK} का उपयोग करके लॉगिन करें। उपयोगकर्ता नाम: {USERNAME} पासवर्ड: {PASSWORD} . DWSS","category":"TRANSACTION","expiryTime":null}'
ippunjabi='{"mobileNumber":"{PHNO}","message":"ਪਿਆਰੇ {USER}, ਤੁਹਾਨੂੰ ਪੰਜਾਬ ਦੀ ਐਮਗ੍ਰਾਮਸੇਵਾ ਐਪਲੀਕੇਸ਼ਨ ਨਾਲ ਜੁੜਨ ਲਈ ਸੱਦਾ ਦਿੱਤਾ ਗਿਆ ਹੈ। ਕਿਰਪਾ ਕਰਕੇ {LINK} ਦੀ ਵਰਤੋਂ ਕਰਕੇ ਲੌਗਇਨ ਕਰੋ। ਲੌਗਿਨ ਆਈ. ਡੀ. {PHNO} ਪਾਸਵਰਡ {PASSWORD}. -DWSS#1007132062282751049","category":"TRANSACTION","expiryTime":null}'
ipenglish='{"mobileNumber":"{PHNO}","message":"Dear {USER}, You\'ve been invited to mGramSeva Application of Punjab. Please login using {LINK}. Username: {USERNAME} Password: {PASSWORD}  DWSS","category":"TRANSACTION","expiryTime":null}'
iphard= '{"mobileNumber":"9980770587","message":"प्रिय  Mani, आपको पंजाब के mGramSeva एप्लिकेशन में आमंत्रित किया गया है। कृपया https://mgramseva-qa.egov.org.in/mgramseva/login का उपयोग करके लॉगिन करें। उपयोगकर्ता नाम: 9980770587 पासवर्ड: eGov@123. EGOVS","category":"TRANSACTION","expiryTime":null}'
def connect():

    accesstoken = accessToken()
    print(accesstoken)
    wrkbk = \
        openpyxl.load_workbook(path)

    sh = wrkbk.active
    if os.path.exists("sms.txt"):
        os.remove("sms.txt")
    file = open("sms.txt", "a", encoding="utf-8")


    for i in range(2, sh.max_row + 1):
        if(sh.cell(i, 4).value!=None):
            createUser(sh, i, accesstoken, file)
    file.close()
    wrkbk.save(path)
    pushSms()

def createUser(sheet, rowIndex, accesstoken,file):
    user = {}
    cityTenantId = sheet.cell(rowIndex, 2).value
    cityTenantId = str(cityTenantId).replace(' ', '')
    cityTenantId = 'as.' + cityTenantId.lower()
    user['tenantId'] = cityTenantId
    user['name'] = sheet.cell(rowIndex, 3).value
    user['mobileNumber'] = sheet.cell(rowIndex, 4).value
    print("Mobile Number:",user['mobileNumber'])
    # user['fatherOrHusbandName'] = sheet.cell(rowIndex, 5).value
    user['gender'] = sheet.cell(rowIndex, 6).value
    user['emailId'] = sheet.cell(rowIndex, 7).value
    dob1=str(sheet.cell(rowIndex, 8).value)
    # print(dob1)
    dob = datetime.datetime.strptime(dob1,'%Y-%m-%d %H:%M:%S')
    user['dob']=datetime.datetime(dob.year,dob.month,dob.day).strftime("%d/%m/%Y")
    roles = []
    role1 = {}
    cityTenantId= sheet.cell(rowIndex, 2).value
    cityTenantId = str(cityTenantId).replace(' ', '')
    cityTenantId='as.'+cityTenantId.lower()
    multiRole= sheet.cell(rowIndex, 10).value
    multiRoleList=multiRole.split(",")
    #print(multiRoleList)
    if len(multiRoleList) > 1:
        for roleCode in multiRoleList :
            roleTemp = {}
            roleTemp['code']=roleCode
            roleTemp['tenantId'] = cityTenantId
            roleTemp['name'] = ' '.join(part.capitalize() for part in roleCode.split('_'))
            #print("in loop"+roleTemp['code'])
            roles.append(roleTemp)
    else :
        role1['code']=multiRole
        role1['tenantId'] = cityTenantId
        role1['name'] = ' '.join(part.capitalize() for part in multiRole.split('_'))
        roles.append(role1)
        #print("from else" + role1['code'])
    role2 = {}
    if sheet.cell(rowIndex, 12).value is not None:
        role2['tenantId'] = cityTenantId
        role2['code'] = sheet.cell(rowIndex, 12).value
        role2['name'] = ' '.join(part.capitalize() for part in sheet.cell(rowIndex, 12).value.split('_'))
        roles.append(role2)
    user['roles'] = roles
    user['password'] = 'mGram@123'
    #'''.join(random.sample(string.ascii_lowercase + string.digits + string.ascii_uppercase + string.punctuation , k=8))
    #print(user['password']);
    # user['type'] = 'EMPLOYEE'
    user['active'] = True
    # user['status'] = 'ACTIVE'
    # user['permanentAddress']='test'
    user['correspondenceAddress']=cityTenantId
    # user['otpReference']='test'
    # user['userName']=sheet.cell(rowIndex, 4).value
    requestInfo = {'authToken': accesstoken}

    userRequest = {}
    userRequest['requestInfo']=requestInfo
    userRequest['user']=user
    employeeSearchResponse= employeeExist(user['mobileNumber'], user['tenantId'], requestInfo)
    if (employeeSearchResponse == None or len(employeeSearchResponse) == 0):
        createEmployeesFieldUser(sheet,rowIndex,accesstoken,user)
    else:
        updateFieldUser(sheet,rowIndex,user,employeeSearchResponse,requestInfo)

def updateFieldUser(sheet, rowIndex, user, employeeSearchResponse,requestInfo):
    add_roles_to_jurisdictions(employeeSearchResponse)
    # Extracting the specific cell values for tenant and roles
    cityTenantId = str(sheet.cell(row=rowIndex, column=2).value).strip().replace(' ', '').lower()
    cityTenantId = f'as.{cityTenantId}'
    roles = str(sheet.cell(row=rowIndex, column=10).value).strip().split(',')

    employees_list = employeeSearchResponse[0]
    current_jurisdictions = employees_list['jurisdictions']

    # Extract current user roles and jurisdictions from the employee search response
    current_roles = user['roles']
    # Update roles in the user
    updated_roles = deepcopy(current_roles)

    # Add new roles from the sheet if not already present
    for role in roles:
        role_name = role.replace('_', ' ').title()
        if not any(r['code'] == role and r['tenantId'] == cityTenantId for r in current_roles):
            updated_roles.append({
                'name': role_name,
                'code': role,
                'tenantId': cityTenantId,
                'description': None
            })

    # Update jurisdictions
    jurisdiction_updated = False
    for jurisdiction in current_jurisdictions:
        if jurisdiction['tenantId'] == cityTenantId:
            jurisdiction_updated = True

            # Merge existing roles with the updated roles
            existing_roles = jurisdiction.get('roles', [])
            combined_roles = {role['code']: role for role in existing_roles}

            # Add new roles to combined_roles if they don't exist
            for role in updated_roles:
                combined_roles[role['code']] = role

            jurisdiction['roles'] = list(combined_roles.values())
            break

    if not jurisdiction_updated:
        # Add a new jurisdiction entry if the tenant ID was not found
        current_jurisdictions.append({
            'hierarchy': cityTenantId,
            'boundaryType': 'City',  # Adjust based on requirements
            'boundary': cityTenantId,
            'tenantId': cityTenantId,
            'roles': updated_roles
        })

    # Update the original user object with updated roles
    current_roles = employeeSearchResponse[0]['user']['roles']
    new_roles = user['roles']
    current_roles.extend(new_roles)

    # Return the updated employee search response
    employeeSearchResponse[0]['jurisdictions'] = current_jurisdictions
    employeeUpdateRequest = {
        "RequestInfo": requestInfo,
        "Employees": employeeSearchResponse
    }
    print("Update Field User Request Body:", employeeUpdateRequest)
    current_time_milliseconds = int(time.time() * 1000)

    update_employee_url = f"{host}egov-hrms/employees/_update?tenantId={cityTenantId}&_={current_time_milliseconds}"

    response = requests.post(update_employee_url, json=employeeUpdateRequest)
    jsondata = response.json()

    if response.status_code == 202:
        employee_data = jsondata.get('Employees', [])[0]
        user_id = employee_data['user']['id']
        print("Successfully updated division employee:", user_id)
        sheet.cell(rowIndex, 15).value = user_id
        sheet.cell(rowIndex, 16).value = "Success(Updated)"
    else:
        error = jsondata['Errors'][0]['code'] + ' ' + jsondata['Errors'][0]['message']
        sheet.cell(rowIndex, 15).value = error
        sheet.cell(rowIndex, 16).value = "Failure(Updated)"
        print("Error updating division employee:", error)

def add_roles_to_jurisdictions(employee_search_response):
    # Extract the roles from the user
    user_roles = employee_search_response[0]['user']['roles']

    # Create a mapping of tenantId to roles
    tenant_roles_map = {}
    for role in user_roles:
        tenant_id = role.get('tenantId')
        if tenant_id:
            if tenant_id not in tenant_roles_map:
                tenant_roles_map[tenant_id] = []
            tenant_roles_map[tenant_id].append(role)

    # Add roles to each jurisdiction based on tenantId
    for jurisdiction in employee_search_response[0]['jurisdictions']:
        tenant_id = jurisdiction.get('tenantId')
        if tenant_id in tenant_roles_map:
            jurisdiction['roles'] = tenant_roles_map[tenant_id]
        else:
            jurisdiction['roles'] = []  # No roles for this tenantId

    return employee_search_response

def createEmployeesFieldUser(sheet, rowIndex, accesstoken, user):
    cityTenantId = sheet.cell(rowIndex, 2).value
    cityTenantId = str(cityTenantId).replace(' ', '')
    cityTenantId = 'as.' + cityTenantId.lower()
    current_time_milliseconds = int(time.time() * 1000)
    user["dob"] = "805075200000"

    jurisdiction_data = []
    tenant_roles_map = {}

    # Ensure role names are properly set
    for role in user['roles']:
        if 'name' not in role:
            role['name'] = role['code']

    roles = user['roles']

    # Organize roles by tenant ID
    for role in roles:
        tenant_id = role["tenantId"]
        if tenant_id not in tenant_roles_map:
            tenant_roles_map[tenant_id] = []
        tenant_roles_map[tenant_id].append(role)

    # Build jurisdiction data
    for tenant_id, tenant_roles in tenant_roles_map.items():
        new_data_entry = {
            "hierarchy": tenant_id,
            "boundaryType": "City",
            "boundary": tenant_id,
            "tenantId": tenant_id,
            "division": {},
            "roles": tenant_roles,
            "divisionBoundary": []  # Division-specific field
        }
        jurisdiction_data.append(new_data_entry)

    create_employee_body = {
        "Employees": [
            {
                "tenantId": cityTenantId,
                "employeeStatus": "EMPLOYED",
                "dateOfAppointment": current_time_milliseconds,
                "employeeType": "PERMANENT",
                "jurisdictions": jurisdiction_data,
                "assignments": [
                    {
                        "fromDate": current_time_milliseconds,
                        "isCurrentAssignment": "true",
                        "department":  sheet.cell(rowIndex, 13).value,
                        "designation": sheet.cell(rowIndex, 14).value
                    }
                ],
                "user": user,
                "serviceHistory": [],
                "education": [],
                "tests": []
            }
        ],
        "key": "CREATE",
        "action": "CREATE",
        "RequestInfo": {
            "apiId": "Rainmaker",
            "authToken": accesstoken,
            "userInfo": {
                "uuid": "7dff7efc-3fcd-4bda-83ac-80bbc69d97d1",
                "tenantId": cityTenantId
            },
            "msgId": "1695277002383|en_IN",
            "plainAccessRequest": {}
        }
    }

    print("Create Field User Request Body:", create_employee_body)

    create_employee_url = f"{host}egov-hrms/employees/_create?tenantId={cityTenantId}&_={current_time_milliseconds}"
    response = requests.post(create_employee_url, json=create_employee_body)
    jsondata = response.json()

    if response.status_code == 202:
        employee_data = jsondata.get('Employees', [])[0]
        user_id = employee_data['user']['id']
        print("Successfully created division employee:", user_id)
        sheet.cell(rowIndex, 15).value = user_id
        sheet.cell(rowIndex, 16).value = "Success"
    else:
        error = jsondata['Errors'][0]['code'] + ' ' + jsondata['Errors'][0]['message']
        sheet.cell(rowIndex, 15).value = error
        sheet.cell(rowIndex, 16).value = "Failure"
        print("Error creating division employee:", error)

def pushSms():
    result = subprocess.run(cmd,shell=True,stdin=subprocess.PIPE)
    print('done');
def addSms(userName,phno,pwd,url,template,file):
    ipTemp = template
    ipTemp = ipTemp.replace("{USER}", userName)
    ipTemp = ipTemp.replace("{PASSWORD}", pwd)
    ipTemp = ipTemp.replace("{PHNO}", str(phno))
    ipTemp = ipTemp.replace("{USERNAME}", str(phno))
    ipTemp = ipTemp.replace("{LINK}", url)
    print(ipTemp);
    file.write(ipTemp)
    file.write("\n")

def employeeExist(phone, tenantId, requestInfo):
    # Construct the employee search request
    employeeSearchRequest = {
        "RequestInfo": requestInfo
    }

    # Define the search URL
    url = f"{host}egov-hrms/employees/_search?tenantId={tenantId}&phone={phone}"

    # Make the POST request to the HRMS API
    post_response = requests.post(
        url=url,
        headers={'Content-type': 'application/json'},
        json=employeeSearchRequest
    )

    # Get the response data as JSON
    jsondata = post_response.json()

    print("Search employee result:",jsondata)

    # Return the employee data if found
    return jsondata.get('Employees')

def accessToken():
    query = {
        'username': username,
        'password': password,
        'userType': 'EMPLOYEE',
        'scope': 'read',
        'grant_type': 'password',
    }
    query['tenantId'] = 'as'
    response = requests.post(host + 'user/oauth/token', data=query,
                             headers={
                                 'Connection': 'keep-alive',
                                 'content-type': 'application/x-www-form-urlencoded',
                                 'origin': host,
                                 'Authorization': 'Basic ZWdvdi11c2VyLWNsaWVudDo=',
                             })
    jsondata = response.json()
    print(jsondata)
    return jsondata.get('access_token')

if __name__ == '__main__':
    connect()