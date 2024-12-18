import openpyxl
import json

json_file = "tenants.json"
# Opening JSON file
f = open(json_file)

# returns JSON object as
# a dictionary
data = json.load(f)
# Iterating through the json
# list
json_data = {

}
for i in data['tenants']:
    json_data[i['code']] = i
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("KA_data_IMIS .xlsx",data_only=True)

# Define variable to read sheet
dataframe1 = dataframe["Sheet1"]

# Iterate the loop to read the cell values
count = 0
change = {
    "tenantId": "as",
    "moduleName": "tenant",
    "tenants": [{
        "code": "as",
        "name": "Assam",
        "description": "Assam",
        "logoId": "https://s3.ap-south-1.amazonaws.com/pb-egov-assets/pb.jalandhar/logo.png",
        "imageId": "https://s3.ap-south-1.amazonaws.com/pb-egov-assets/pb.jalandhar/logo.png",
        "domainUrl": "www.mcjalandhar.in",
        "type": "CITY",
        "twitterUrl": "https://twitter.com/search?q=%23jalandhar",
        "facebookUrl": "https://www.facebook.com/city/jalandhar-Punjab",
        "emailId": "complaints.mcj@gmail.com",
        "OfficeTimings": {
            "Mon - Fri": "9.00 AM - 6.00 PM"
        },
        "city": {
            "name": "Assam State",
            "localName": None,
            "districtCode": None,
            "districtName": None,
            "regionName": None,
            "ulbGrade": "ST",
            "longitude": 75.5761829,
            "latitude": 31.3260152,
            "shapeFileLocation": None,
            "captcha": None,
            "code": "15",
            "ddrName": "Assam",
            "projectId": "15"
        },
        "address": "Assam-144001",
        "pincode": [],
        "contactNumber": "0181-2227015",
        "pdfHeader": "PB_PDF_HEADER",
        "pdfContactDetails": "PB_CONTACT_DETAILS"
    }
    ]
}
changed_v_code = []


def create_tenant(code, gpwsc, scheme, district_code, district_name, block_code,block_name,panchayat_code,panchayat_name,village_code,village_name):
    village = {}
    city_tenant_id = 'as.' + scheme.lower()
    village['code'] = city_tenant_id.replace(" ", "")
    village['name'] = gpwsc
    village['description'] = gpwsc
    village['logoId'] = ""
    village['imageId'] = None
    village['domainUrl'] = ""
    village['city'] = "CITY"
    village['twitterUrl'] = None
    village['facebookUrl'] = None
    village['emailId'] = ""
    village['OfficeTimings'] = {"Mon - Fri": "9.00 AM - 6.00 PM"}
    city = {'name': gpwsc, 'localName': gpwsc, 'districtCode': district_code, 'districtName': district_name,
            'blockcode': block_code,
            'blockname':block_name,
            'panchayatcode':panchayat_code,
            'panchayatname':panchayat_name,
            'villageCode':village_code,
            'villageName':village_name,
            'ulbGrade': "", 'longitude': None, 'latitude': None, 'captcha': None, 'shapeFileLocation': None,
            'code': str(code), 'ddrName': gpwsc, 'projectId': str(code)}
    village['city'] = city
    village['address'] = gpwsc
    village['districtCode']= district_code
    village['districtName']= district_name
    village['blockcode']= block_code
    village['blockname']= block_name
    village['panchayatcode']= panchayat_code
    village['panchayatname']= panchayat_name
    village['villageCode']= village_code
    village['villageName']= village_name
    village['pincode'] = []
    village['contactNumber'] = ""
    village['pdfHeader'] = ""
    village['pdfContactDetails'] = ""
    return village

for row in range(0, 1):
    arr = []
    for col in dataframe1.iter_cols(0, 6):
        arr.append(str(col[row].value))
    for i in arr:
        print(str(arr.index(i)), ":", i, end=", ")
    print("", end="\n")
print("----------------------------------------------------------------------------------------------", end="\n")
for row in range(1, 2): #specify range
    arr = []
    for col in dataframe1.iter_cols(0, 11):
        arr.append(str(col[row].value))
    tenant = json_data.get("as." + arr[0])
    if tenant is None:
        tenant = create_tenant(arr[8], arr[1], arr[0], arr[2], arr[3], arr[4], arr[5], arr[6], arr[7],arr[9],arr[10])
    for i in arr:
        print(str(arr.index(i)), ":", i, end=", ")
    tenant["city"]["districtCode"] = arr[2]
    tenant["city"]["districtName"] = arr[3]
    tenant["city"]["regionName"] = arr[4]
    if tenant["city"]["code"] != arr[5]:
        changed_v_code.append(tenant)
    change["tenants"].append(tenant)
    print("", end="\n")
    count += 1
if len(changed_v_code) == 0:
    print("No Mismatch", end="\n")
else:
    print("mismatch villages: ", changed_v_code)
print("count: " + str(count))
print(change)
# Serializing json
json_object = json.dumps(change, indent=4)

# Writing to sample.json
with open("tenants_new.json", "w") as outfile:
    outfile.write(json_object)
# Closing file
f.close()
