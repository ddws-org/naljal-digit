import concurrent
# !/usr/bin/python
# -*- coding: utf-8 -*-
import csv
import pandas as pd
import numpy as np
import pytz
import requests
import json
import openpyxl
import time
import datetime
import subprocess
import os
import string
import random
import re
import concurrent.futures
import pytz

host = 'https://naljalseva.jjm.gov.in/uat/'
loginUrl = host + 'mgramseva/login'
username = ''  # superuser mobile no
password = ''  # superuser password
stateTenantId = 'as'
path = '' # file path
oldReadingDate = '2024-09-01'



def connect():
    accesstokenjson = accessToken()
    accesstoken = accesstokenjson.get('access_token')
    userInfo = accesstokenjson.get('UserRequest')
    print(userInfo)

    wrkbk = openpyxl.load_workbook(path)
    sh = wrkbk.active

    if os.path.exists("result.txt"):
        os.remove("result.txt")
    file = open("result.txt", "a")

    # Create a thread pool to process each row concurrently
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:  # You can adjust max_workers as needed
        futures = []

        for i in range(2, sh.max_row + 1):
            # Submit each row to the thread pool
            futures.append(executor.submit(process_row, i, sh, accesstoken, userInfo, file, oldReadingDate))

        # Wait for all threads to complete
        for future in concurrent.futures.as_completed(futures):
            future.result()  # This will also raise exceptions if any were thrown in threads

    file.close()

    wrkbk.save(path)
    print('done')


def isPhoneNumberValid(s):
    Pattern = re.compile("(0|91)?[1-9][0-9]{9}")
    if (Pattern.match(str(s)) and (len(str(s)) == 10)):
        return True
    else:
        return False


def process_row(i, sheet, accesstoken, userInfo, file, oldReadingDate):
    status = sheet.cell(i, 60000).value
    if (True):
        cityTenantId = str(sheet.cell(i, 14).value)
        propertySearchResponse = propertyExist('as.' + cityTenantId.lower(), sheet.cell(i, 6).value, accesstoken,
                                               userInfo)
        if (propertySearchResponse != None):
            if len(propertySearchResponse) == 0:
                createProperty(sheet, i, accesstoken, file, userInfo)
            else:
                previousReadingDate = datetime.datetime.strptime(oldReadingDate, '%Y-%m-%d')
                print(previousReadingDate)
                datevalue = datetime.datetime(previousReadingDate.year, previousReadingDate.month,
                                              previousReadingDate.day).timestamp()
                print("property already present")
                sheet.cell(i, 23).value = 'Success'
                sheet.cell(i, 24).value = 'property already present'


def propertyExist(tenantId, oldConnectionNo, accesstoken, userInfo):
    propertySearchRequest = {}
    requestInfo = {'authToken': accesstoken, 'userInfo': userInfo, 'apiId': 'mgramseva',
                   'ver': 1,
                   'ts': "",
                   'action': '_search',
                   'did': 1,
                   'key': "",
                   'msgId': '20170310130900|en_IN'}

    propertySearchRequest['RequestInfo'] = requestInfo
    print(propertySearchRequest)
    propertySearchRequest['userType'] = 'EMPLOYEE'
    post_response = requests.post(url="http://ws-services.uat:8080/uat/"
                                      + 'ws-services/wc/_search?tenantId=' + tenantId + '&oldConnectionNumber=' + str(
        oldConnectionNo),
                                  headers={'Content-type': 'application/json'},
                                  json=propertySearchRequest)

    if (post_response == None):
        print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
        return None
    else:
        try:
            jsondata = post_response.json()
            print("search ws result");
            print(jsondata);
            return jsondata.get('WaterConnection')
        except:
            print("some exception--->>"+post_response.json());
            print(post_response.json());
            return None


def propertyExist2(tenantId, ptid, accesstoken, userInfo):
    propertySearchRequest = {}
    requestInfo = {'authToken': accesstoken, 'userInfo': userInfo, 'apiId': 'mgramseva',
                   'ver': 1,
                   'ts': "",
                   'action': '_search',
                   'did': 1,
                   'key': "",
                   'msgId': '20170310130900|en_IN'}

    propertySearchRequest['RequestInfo'] = requestInfo
    propertySearchRequest['propertyCriteria'] = {
        'tenantId': tenantId,
        'propertyIds': [
            ptid
        ]
    }
    print(propertySearchRequest)
    propertySearchRequest['userType'] = 'EMPLOYEE'

    post_response = requests.post(url="http://property-services.uat:8080/uat/"
                                      + 'property-services/property/_search',
                                  headers={'Content-type': 'application/json'},
                                  json=propertySearchRequest)
    print("search property body", propertySearchRequest)
    if (post_response == None):
        print("no response from propertyExist api call" + ptid + " " + tenantId.upper());
        return None
    else:
        try:
            jsondata = post_response.json()
            print("search property result");
            print(jsondata);
            return jsondata
        except:
            print("some exception");
            print(post_response.json());
            return None


def createProperty(sheet, rowIndex, accesstoken, file, userInfo):
    print("create Prop")
    owners = []
    owner = {}
    cityTenantId = str(sheet.cell(rowIndex, 14).value)
    print(cityTenantId)
    cityTenantId = 'as.' + str(cityTenantId).lower()
    # owner['mobileNumber'] = sheet.cell(rowIndex, 5).value
    # mobileNumber = sheet.cell(rowIndex, 5).value
    # val = isPhoneNumberValid(mobileNumber)
    val = True
    if val == False:
        print("Property Creation failed ! ")
        sheet.cell(rowIndex, 23).value = 'Failure'
        error = "Invalid Mobile Number"
        sheet.cell(rowIndex, 24).value = error
    else:
        num = sheet.cell(rowIndex, 5).value
        if num == 'NONE':
            num = None
        else:
            if (isPhoneNumberValid(num) == False):
                num = None
        owner['mobileNumber'] = num
        owner['name'] = str(sheet.cell(rowIndex, 2).value)
        print(owner['name'])
        owner['gender'] = str(sheet.cell(rowIndex, 3).value).upper()
        fatherOrHusbandName = str(sheet.cell(rowIndex, 4).value).upper()
        owner['fatherOrHusbandName'] = fatherOrHusbandName
        owner['ownerType'] = 'NONE'
        # owner['emailId'] ='null'
        owners.append(owner)
        if (sheet.cell(rowIndex, 24).value == None):
            Property = {}
            print("creating property...")
            Property['tenantId'] = cityTenantId
            Property['ownershipCategory'] = 'INDIVIDUAL'
            locality = {}
            address1 = {}
            address1['street'] = str(sheet.cell(rowIndex, 10).value)
            # address1['geoLocation'] = cityTenantId
            address1['doorNo'] = str(sheet.cell(rowIndex, 9).value)
            address1['landmark'] = 'null'
            # locality['code']=sheet.cell(rowIndex, 11).value
            locality['code'] = 'WARD1'
            locality['area'] = 'null'
            address1['locality'] = locality
            Property['address'] = address1

            Property['owners'] = owners
            prop = str(sheet.cell(rowIndex, 7).value).upper()
            print(prop)
            if (prop == 'NONE'):
                prop = 'RESIDENTIAL'
            Property['propertyType'] = prop
            # Property['propertyType']=Property['propertyType'].upper()
            Property['noOfFloors'] = 1
            Property['usageCategory'] = prop
            Property['landArea'] = 1
            Property['creationReason'] = 'CREATE'
            Property['source'] = 'MUNICIPAL_RECORDS'
            Property['channel'] = 'CITIZEN'
            requestInfo = {'authToken': accesstoken, 'userInfo': userInfo, 'apiId': 'mgramseva',
                   'ver': 1,
                   'ts': "",
                   'action': '_search',
                   'did': 1,
                   'key': "",
                   'msgId': '20170310130900|en_IN'}

            ptRequest = {}
            ptRequest['RequestInfo'] = requestInfo
            ptRequest['Property'] = Property
            print("ptRequest!!")
            print(ptRequest)
            post_response = requests.post(url="http://property-services.uat:8080/uat/"
                                              + 'property-services/property/_create',
                                          headers={'Content-type': 'application/json'},
                                          json=ptRequest)
            if (post_response == None):
                print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
                return None
            else:
                try:
                    jsondata = post_response.json()
                    print(jsondata)
                    jsondata.get('Properties') == None
                    if (jsondata.get('Properties') != None):
                        ptid = jsondata.get('Properties')[0]['propertyId']
                        print("Property Created . with Id ")
                        print(ptid)
                        owner['mobileNumber'] = jsondata.get('Properties')[0]['owners'][0]['mobileNumber']
                        sheet.cell(rowIndex, 24).value = ptid
                        createWaterConnection(sheet, rowIndex, accesstoken, file, ptid, cityTenantId, owner,userInfo)
                    else:
                        print("Property Creation failed ! ")
                        # print(jsondata);
                        sheet.cell(rowIndex, 23).value = 'Failure'
                        error = jsondata['Errors'][0]['code'] + ' ' + jsondata['Errors'][0]['message']
                        sheet.cell(rowIndex, 24).value = error
                except:
                    print("some exception");
                    print(post_response.json());
                    return None

        else:
            print("property already present")
            ptid = sheet.cell(rowIndex, 18).value
            wc = propertyExist2(cityTenantId, ptid, accesstoken, userInfo)
            wa = propertyExist(cityTenantId, sheet.cell(rowIndex, 6).value, accesstoken, userInfo)
            if (wc.get('Properties') != None):
                ptid = wc.get('Properties')[0]['propertyId']
                print("Property Created . with Id ")
                print(ptid)
                owner['mobileNumber'] = wc.get('Properties')[0]['owners'][0]['mobileNumber']
                sheet.cell(rowIndex, 18).value = ptid
                if (wa.get('WaterConnection') != None):
                    if (len(wa.get('WaterConnection')) == 0):
                        createWaterConnection(sheet, rowIndex, accesstoken, file, ptid, cityTenantId, owner,userInfo)
                    else:
                        # oldReadingDate = str(sh.cell(i, 13).value)
                        previousReadingDate = datetime.datetime.strptime(oldReadingDate, '%Y-%m-%d')
                        print(previousReadingDate)
                        datevalue = datetime.datetime(previousReadingDate.year, previousReadingDate.month,
                                                      previousReadingDate.day).timestamp()
                        # print(datevalue * 1000)
                        print("property already present")
                        sh.cell(i, 23).value = 'Success'
                        sh.cell(i, 24).value = 'property already present'
            else:
                print("Property Creation failed ! ")
                # print(jsondata);
                sheet.cell(rowIndex, 23).value = 'Failure'
                error = wc['Errors'][0]['code'] + ' ' + wc['Errors'][0]['message']
                sheet.cell(rowIndex, 24).value = error


def createWaterConnection(sheet, rowIndex, accesstoken, file, propertyId, tenantId, owner,userInfo):
    print('creating connection.....');
    WaterConnection = {}
    WaterConnection['propertyId'] = propertyId
    # WaterConnection['applicationNo'] =1222
    WaterConnection['tenantId'] = tenantId;
    WaterConnection['action'] = 'SUBMIT'
    WaterConnection['meterInstallationDate'] = 111111111111
    # WaterConnection['documents'] =
    WaterConnection['proposedTaps'] = 1
    WaterConnection['noOfTaps'] = 1
    WaterConnection['arrears'] = sheet.cell(rowIndex, 15).value
    connType = sheet.cell(rowIndex, 8).value
    if (connType == None):
        connType = 'Non_Metered'
    WaterConnection['connectionType'] = connType
    if (connType == 'Metered'):
        WaterConnection['oldConnectionNo'] = sheet.cell(rowIndex, 6).value
        WaterConnection['meterId'] = sheet.cell(rowIndex, 12).value
        # oldReadingDate = str(sheet.cell(rowIndex, 13).value)
        previousReadingDate = datetime.datetime.strptime(oldReadingDate, '%Y-%m-%d')
        datevalue = datetime.datetime(previousReadingDate.year, previousReadingDate.month,
                                      previousReadingDate.day, 0, 0, 0).timestamp()
        WaterConnection['previousReadingDate'] = datevalue * 1000
        WaterConnection['previousReading'] = sheet.cell(rowIndex, 16).value
    else:
        WaterConnection['oldConnectionNo'] = sheet.cell(rowIndex, 6).value
        # oldReadingDate = str(sheet.cell(rowIndex, 13).value)
        previousReadingDate = datetime.datetime.strptime(oldReadingDate, '%Y-%m-%d')
        datevalue = datetime.datetime(previousReadingDate.year, previousReadingDate.month,
                                      previousReadingDate.day, 0, 0, 0).timestamp()
        WaterConnection['previousReadingDate'] = int(datevalue * 1000)
        print(int(datevalue * 1000))

    prop = str(sheet.cell(rowIndex, 7).value).upper()
    if (prop == 'NONE'):
        prop = 'RESIDENTIAL'
    WaterConnection['propertyType'] = prop
    WaterConnection['proposedPipeSize'] = 1
    if (str(sheet.cell(rowIndex, 17).value) != 'NONE'):
        WaterConnection['imisNumber'] = sheet.cell(rowIndex, 17).value
    if (str(sheet.cell(rowIndex, 18).value) != 'NONE'):
        WaterConnection['schemeId'] = sheet.cell(rowIndex, 18).value
    if (str(sheet.cell(rowIndex, 19).value) != 'NONE'):
        WaterConnection['schemeName'] = sheet.cell(rowIndex, 19).value
    if (str(sheet.cell(rowIndex, 20).value) != 'NONE'):
        WaterConnection['ihlDetail'] = sheet.cell(rowIndex, 20).value
    if (str(sheet.cell(rowIndex, 21).value) != 'NONE'):
        WaterConnection['sbmAccountno'] = sheet.cell(rowIndex, 21).value
    if (str(sheet.cell(rowIndex, 25).value) != 'NONE'):
        WaterConnection['villageId'] = sheet.cell(rowIndex, 25).value
    connectionHolders = []
    connectionHolders.append(owner)
    WaterConnection['connectionHolders'] = connectionHolders
    additionalDetails = {}
    # additionalDetails['initialMeterReading'] =sheet.cell(rowIndex, 16).value
    additionalDetails['initialMeterReading'] = 0
    # additionalDetails['meterReading'] =sheet.cell(rowIndex, 13).value
    # additionalDetails['locality'] =sheet.cell(rowIndex, 11).value
    additionalDetails['locality'] = 'WARD1'
    if (str(sheet.cell(rowIndex, 26).value) != 'NONE'):
        additionalDetails['villageName'] = sheet.cell(rowIndex, 26).value
    additionalDetails['propertyType'] = prop
    additionalDetails['street'] = sheet.cell(rowIndex, 10).value
    additionalDetails['doorNo'] = sheet.cell(rowIndex, 9).value
    # fix hardcoding
    additionalDetails['category'] = 'APL'
    additionalDetails['subCategory'] = 'GENERAL'
    # additionalDetails['collectionAmount'] =
    # additionalDetails['action'] =
    WaterConnection['additionalDetails'] = additionalDetails
    processInstance = {}
    processInstance['action'] = 'SUBMIT'
    WaterConnection['processInstance'] = processInstance

    requestInfo = {'authToken': accesstoken, 'userInfo': userInfo, 'apiId': 'mgramseva',
                   'ver': 1,
                   'ts': "",
                   'action': '_search',
                   'did': 1,
                   'key': "",
                   'msgId': '20170310130900|en_IN'}
    waterRequest = {}
    waterRequest['RequestInfo'] = requestInfo
    waterRequest['WaterConnection'] = WaterConnection

    # print("Search Result :\n" )
    # print(userSearchResponse);

    print("water request!!!")
    print(waterRequest)
    post_response = requests.post(url="http://ws-services.uat:8080/uat/"
                                      + 'ws-services/wc/_create',
                                  headers={'Content-type': 'application/json'},
                                  json=waterRequest)
    print(post_response.json())
    if (post_response == None):
        print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
        return None
    else:
        try:
            jsondata = post_response.json()
            if (jsondata['ResponseInfo'] != None):
                # print(jsondata);
                print("connection created with Id")
                print(jsondata['WaterConnection'][0]['connectionNo']);
                sheet.cell(rowIndex, 22).value = jsondata['WaterConnection'][0]['connectionNo'];
                sheet.cell(rowIndex, 23).value = 'Success'
            else:
                # print(jsondata);
                sheet.cell(rowIndex, 23).value = 'Failure'
                error = jsondata['Errors'][0]['code'] + ' ' + jsondata['Errors'][0]['message']
                # error = error +' ' jsondata['Errors'][0]['message']
                sheet.cell(rowIndex, 24).value = error
        except:
            print("some exception");
            print(post_response.json());
            return None


def accessToken():
    query = {
        'username': username,
        'password': password,
        'userType': 'EMPLOYEE',
        'scope': 'read',
        'grant_type': 'password',
    }
    query['tenantId'] = 'as'
    response = requests.post(host + 'user/oauth/token', data=query,
                             headers={
                                 'Connection': 'keep-alive',
                                 'content-type': 'application/x-www-form-urlencoded',
                                 'origin': host,
                                 'Authorization': 'Basic ZWdvdi11c2VyLWNsaWVudDo=',
                             })
    if (response == None):
        print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
        return None
    else:
        try:
            jsondata = response.json()
            print(jsondata)
            return jsondata
        except:
            print("some exception");
            print(response.json());
            return None


if __name__ == '__main__':
    tzInfo = pytz.timezone('Asia/Kolkata')
    time = datetime.datetime.now(tz=tzInfo)
    print("start_time -->", time)
    connect()
    time_end = datetime.datetime.now(tz=tzInfo)
    print("end_time -->", time_end)
