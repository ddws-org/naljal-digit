
#!/usr/bin/python
# -*- coding: utf-8 -*-
import csv
import pandas as pd
import numpy as np
import pytz
import requests
import json
import openpyxl
import time
import datetime
import subprocess
import os
import string
import random
import re

host = 'https://mgramseva-uat.psegs.in/'
loginUrl = host + 'mgramseva/login'
username = '9845530047' #9845530047
password = 'eGov@123'
stateTenantId='pb'
path='consumer-master-6-dwss.xlsx'
def connect():

    accesstokenjson = accessToken()
    accesstoken = accesstokenjson.get('access_token')
    userInfo = accesstokenjson.get('UserRequest')
    print(userInfo)
    wrkbk = \
        openpyxl.load_workbook(path)
                               
    sh = wrkbk.active
    if os.path.exists("result.txt"):
        os.remove("result.txt")
    file = open("result.txt", "a")

    for i in range(2, sh.max_row + 1):
        status = sh.cell(i, 17).value
        if (status == None or status == 'Failure'):
            cityTenantId = sh.cell(i, 14).value
            propertySearchResponse = propertyExist('pb.'+cityTenantId.lower(), sh.cell(i, 6).value, accesstoken,userInfo)
            if (propertySearchResponse != None):
                if(len(propertySearchResponse) == 0):
                    createProperty(sh, i, accesstoken, file)
                else:
                    oldReadingDate = str(sh.cell(i, 13).value)
                    previousReadingDate = datetime.datetime.strptime(oldReadingDate, '%Y-%m-%d 00:00:00')
                    print(previousReadingDate)
                    datevalue = datetime.datetime(previousReadingDate.year, previousReadingDate.month,
                                                  previousReadingDate.day).timestamp()
                    # print(datevalue * 1000)
                    print("property already present")
                    sh.cell(i, 17).value = 'Success'
                    sh.cell(i, 20).value = 'property already present'

    file.close()

    wrkbk.save(path)
    print('done')

def isPhoneNumberValid(s):
    Pattern = re.compile("(0|91)?[6-9][0-9]{9}")
    if(Pattern.match(str(s)) and (len(str(s))==10)):
        return True
    else:
        return False

def propertyExist(tenantId, oldConnectionNo, accesstoken, userInfo):
    propertySearchRequest = {}
    requestInfo = {'authToken': accesstoken, 'userInfo': userInfo,'apiId': 'mgramseva',
        'ver': 1,
        'ts': "",
        'action': '_search',
        'did': 1,
        'key': "",
        'msgId': '20170310130900|en_IN'}


    propertySearchRequest['RequestInfo'] = requestInfo
    print(propertySearchRequest)
    propertySearchRequest['userType'] = 'EMPLOYEE'
    post_response = requests.post(url=host
                                      + 'ws-services/wc/_search?tenantId='+tenantId+'&oldConnectionNumber='+str(oldConnectionNo),
                                  headers={'Content-type': 'application/json'},
                                  json=propertySearchRequest)

    if(post_response == None):
        print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
        return None
    else:
        try:
            jsondata = post_response.json()
            print("search ws result");
            print(jsondata);
            return jsondata.get('WaterConnection')
        except:
            print("some exception");
            print(post_response);
            return None

def createProperty(sheet, rowIndex, accesstoken,file):
    owners = []
    owner = {}
    cityTenantId = sheet.cell(rowIndex, 14).value
    print(cityTenantId)
    cityTenantId = 'pb.'+str(cityTenantId).lower()
    # owner['mobileNumber'] = sheet.cell(rowIndex, 5).value
    mobileNumber = sheet.cell(rowIndex, 5).value
    val = isPhoneNumberValid(mobileNumber)
    if(val == False):
        print("Property Creation failed ! ")
        sheet.cell(rowIndex, 17).value = 'Failure'
        error = "Invalid Mobile Number"
        sheet.cell(rowIndex, 20).value = error
    else:
        owner['mobileNumber'] = sheet.cell(rowIndex, 5).value
        owner['name'] = sheet.cell(rowIndex, 2).value
        print(owner['name'])
        owner['gender'] = sheet.cell(rowIndex, 3).value
        owner['fatherOrHusbandName'] = sheet.cell(rowIndex, 4).value
        owner['ownerType'] = 'NONE'
        # owner['emailId'] ='null'
        owners.append(owner)
        if(sheet.cell(rowIndex, 18).value == None ):
            Property = {}
            print("creating property...")
            Property['tenantId'] = cityTenantId
            Property['ownershipCategory'] = 'INDIVIDUAL'
            locality = {}
            address1 = {}
            address1['street']=sheet.cell(rowIndex, 10).value
            #address1['geoLocation'] = cityTenantId
            address1['doorNo']=sheet.cell(rowIndex, 9).value
            address1['landmark'] = 'null'
            # locality['code']=sheet.cell(rowIndex, 11).value
            locality['code']='WARD1'
            locality['area'] ='null'
            address1['locality']=locality
            Property['address'] = address1


            Property['owners'] = owners
            Property['propertyType'] = str(sheet.cell(rowIndex, 7).value).upper()
            #Property['propertyType']=Property['propertyType'].upper()
            Property['noOfFloors'] = 1
            Property['usageCategory'] = str(sheet.cell(rowIndex, 7).value).upper()
            Property['landArea'] = 1
            Property['creationReason'] = 'CREATE'
            Property['source'] = 'MUNICIPAL_RECORDS'
            Property['channel'] = 'CITIZEN'
            requestInfo = {'authToken': accesstoken}

            ptRequest = {}
            ptRequest['RequestInfo']=requestInfo
            ptRequest['Property']=Property

            post_response = requests.post(url=host
                                      + 'property-services/property/_create',
                                      headers={'Content-type': 'application/json'},
                                      json=ptRequest)
            if (post_response == None):
                print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
                return None
            else:
                try:
                    jsondata = post_response.json()
                    if (jsondata.get('Properties') != None):
                        ptid = jsondata.get('Properties')[0]['propertyId']
                        print("Property Created . with Id ")
                        print(ptid)
                        sheet.cell(rowIndex, 18).value = ptid
                        createWaterConnection(sheet, rowIndex, accesstoken, file, ptid, cityTenantId, owner)
                    else:
                        print("Property Creation failed ! ")
                        # print(jsondata);
                        sheet.cell(rowIndex, 17).value = 'Failure'
                        error = jsondata['Errors'][0]['code'] + ' ' + jsondata['Errors'][0]['message']
                        sheet.cell(rowIndex, 20).value = error
                except:
                    print("some exception");
                    print(post_response);
                    return None

        else:
            print("property already present")
            ptid=sheet.cell(rowIndex, 18).value
            createWaterConnection(sheet, rowIndex, accesstoken, file, ptid, cityTenantId, owner)

def createWaterConnection(sheet,rowIndex,accesstoken,file,propertyId,tenantId,owner):
    print('creating connection.....');
    WaterConnection={}
    WaterConnection['propertyId'] =propertyId
    #WaterConnection['applicationNo'] =1222
    WaterConnection['tenantId'] =tenantId;
    WaterConnection['action'] ='SUBMIT'
    WaterConnection['meterInstallationDate'] =111111111111
    #WaterConnection['documents'] =
    WaterConnection['proposedTaps'] =1
    WaterConnection['noOfTaps'] =1
    WaterConnection['arrears'] =sheet.cell(rowIndex, 15).value
    connType=sheet.cell(rowIndex, 8).value
    WaterConnection['connectionType'] =connType
    if(connType=='Metered'):
        WaterConnection['oldConnectionNo'] =sheet.cell(rowIndex, 6).value
        WaterConnection['meterId'] =sheet.cell(rowIndex, 12).value
        oldReadingDate = str(sheet.cell(rowIndex, 13).value)
        previousReadingDate = datetime.datetime.strptime(oldReadingDate, '%Y-%m-%d 00:00:00')
        datevalue = datetime.datetime(previousReadingDate.year, previousReadingDate.month,
                                      previousReadingDate.day,0,0,0).timestamp()
        WaterConnection['previousReadingDate'] = datevalue * 1000
        WaterConnection['previousReading'] =sheet.cell(rowIndex, 16).value
    else:
        WaterConnection['oldConnectionNo'] =sheet.cell(rowIndex, 6).value
        oldReadingDate = str(sheet.cell(rowIndex, 13).value)
        previousReadingDate = datetime.datetime.strptime(oldReadingDate, '%Y-%m-%d 00:00:00')
        datevalue = datetime.datetime(previousReadingDate.year, previousReadingDate.month,
                                      previousReadingDate.day,0,0,0).timestamp()
        WaterConnection['previousReadingDate'] = int(datevalue * 1000)
        print(int(datevalue * 1000))

    WaterConnection['propertyType'] = sheet.cell(rowIndex, 7).value
    WaterConnection['proposedPipeSize'] =1
    connectionHolders=[]
    connectionHolders.append(owner)
    WaterConnection['connectionHolders']=connectionHolders
    additionalDetails={}
    # additionalDetails['initialMeterReading'] =sheet.cell(rowIndex, 16).value
    additionalDetails['initialMeterReading'] =0
    #additionalDetails['meterReading'] =sheet.cell(rowIndex, 13).value
    # additionalDetails['locality'] =sheet.cell(rowIndex, 11).value
    additionalDetails['locality'] ='WARD1'
    additionalDetails['propertyType'] =sheet.cell(rowIndex, 7).value
    additionalDetails['street'] =sheet.cell(rowIndex, 10).value
    additionalDetails['doorNo'] =sheet.cell(rowIndex, 9).value
    #fix hardcoding
    additionalDetails['category'] ='APL'
    additionalDetails['subCategory'] ='GENERAL'
    #additionalDetails['collectionAmount'] =
    #additionalDetails['action'] =
    WaterConnection['additionalDetails'] = additionalDetails
    processInstance={}
    processInstance['action']='SUBMIT'
    WaterConnection['processInstance']=processInstance


    requestInfo = {'authToken': accesstoken}
    waterRequest = {}
    waterRequest['requestInfo']=requestInfo
    waterRequest['WaterConnection']=WaterConnection


    #print("Search Result :\n" )
    #print(userSearchResponse);


    post_response = requests.post(url=host
                              + 'ws-services/wc/_create',
                              headers={'Content-type': 'application/json'},
                              json=waterRequest)
    if (post_response == None):
        print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
        return None
    else:
        try:
            jsondata = post_response.json()
            if (jsondata['ResponseInfo'] != None):
                # print(jsondata);
                print("connection created with Id")
                print(jsondata['WaterConnection'][0]['connectionNo']);
                sheet.cell(rowIndex, 19).value = jsondata['WaterConnection'][0]['connectionNo'];
                sheet.cell(rowIndex, 17).value = 'Success'
            else:
                # print(jsondata);
                sheet.cell(rowIndex, 17).value = 'Failure'
                error = jsondata['Errors'][0]['code'] + ' ' + jsondata['Errors'][0]['message']
                # error = error +' ' jsondata['Errors'][0]['message']
                sheet.cell(rowIndex, 20).value = error
        except:
            print("some exception");
            print(post_response);
            return None

def accessToken():
    query = {
        'username': username,
        'password': password,
        'userType': 'EMPLOYEE',
        'scope': 'read',
        'grant_type': 'password',
        }
    query['tenantId'] = 'pb'
    response = requests.post(host + 'user/oauth/token', data=query,
                             headers={
        'Connection': 'keep-alive',
        'content-type': 'application/x-www-form-urlencoded',
        'origin': host,
        'Authorization': 'Basic ZWdvdi11c2VyLWNsaWVudDo=',
        })
    if (response == None):
        print("no response from propertyExist api call" + oldConnectionNo + " " + tenantId.upper());
        return None
    else:
        try:
            jsondata = response.json()
            print(jsondata)
            return jsondata
        except:
            print("some exception");
            print(response);
            return None



if __name__ == '__main__':
    connect()
