
#!/usr/bin/python
# -*- coding: utf-8 -*-
import csv
import pandas as pd
import numpy as np
import requests
import json
import openpyxl
import time
import datetime
import subprocess
import os
import string
import random

host = 'https://mgramseva-dwss.punjab.gov.in/'
loginUrl = host + 'mgramseva/login'
username = '9448897072' # '9837872921'
password = 'eGov@123'
stateTenantId='pb'
path='demo-user.xlsx'
#cmd=['kubectl', '-it', 'exec', 'kafka-v2-0',  '-n', 'mgramseva', '--', 'kafka-console-producer', '--broker-list', 'localhost:9092', '--topic',  'egov.core.notification.sms']
cmd = "kubectl -it exec kafka-v2-0  -n mgramseva -- kafka-console-producer --broker-list localhost:9092 --topic  egov.core.notification.sms <sms.txt"
iphindi='{"mobileNumber":"{PHNO}","message":"प्रिय {USER}, आपको पंजाब के mGramSeva एप्लिकेशन में आमंत्रित किया गया है। कृपया {LINK} का उपयोग करके लॉगिन करें। उपयोगकर्ता नाम: {USERNAME} पासवर्ड: {PASSWORD} . DWSS","category":"TRANSACTION","expiryTime":null}'
ippunjabi='{"mobileNumber":"{PHNO}","message":"ਪਿਆਰੇ {USER}, ਤੁਹਾਨੂੰ ਪੰਜਾਬ ਦੀ ਐਮਗ੍ਰਾਮਸੇਵਾ ਐਪਲੀਕੇਸ਼ਨ ਨਾਲ ਜੁੜਨ ਲਈ ਸੱਦਾ ਦਿੱਤਾ ਗਿਆ ਹੈ। ਕਿਰਪਾ ਕਰਕੇ {LINK} ਦੀ ਵਰਤੋਂ ਕਰਕੇ ਲੌਗਇਨ ਕਰੋ। ਲੌਗਿਨ ਆਈ. ਡੀ. {PHNO} ਪਾਸਵਰਡ {PASSWORD}. -DWSS#1007132062282751049","category":"TRANSACTION","expiryTime":null}'
ipenglish='{"mobileNumber":"{PHNO}","message":"Dear {USER}, You\'ve been invited to mGramSeva Application of Punjab. Please login using {LINK}. Username: {USERNAME} Password: {PASSWORD}  DWSS","category":"TRANSACTION","expiryTime":null}'
iphard= '{"mobileNumber":"9980770587","message":"प्रिय  Mani, आपको पंजाब के mGramSeva एप्लिकेशन में आमंत्रित किया गया है। कृपया https://mgramseva-qa.egov.org.in/mgramseva/login का उपयोग करके लॉगिन करें। उपयोगकर्ता नाम: 9980770587 पासवर्ड: eGov@123. EGOVS","category":"TRANSACTION","expiryTime":null}'
def connect():

    accesstoken = accessToken()
    print(accesstoken)
    wrkbk = \
        openpyxl.load_workbook(path)
                               
    sh = wrkbk.active
    if os.path.exists("sms.txt"):
        os.remove("sms.txt")
    file = open("sms.txt", "a", encoding="utf-8")


    for i in range(2, sh.max_row + 1):
        if(sh.cell(i, 4).value!=None):
            createUser(sh, i, accesstoken, file)
    file.close()
    wrkbk.save(path)
    pushSms()

def createUser(sheet, rowIndex, accesstoken,file):
    user = {}
    user['tenantId'] = stateTenantId
    user['name'] = sheet.cell(rowIndex, 3).value
    user['mobileNumber'] = sheet.cell(rowIndex, 4).value
    user['fatherOrHusbandName'] = sheet.cell(rowIndex, 5).value
    user['gender'] = sheet.cell(rowIndex, 6).value
    user['emailId'] = sheet.cell(rowIndex, 7).value
    dob1=str(sheet.cell(rowIndex, 8).value)
    print(dob1)
    dob = datetime.datetime.strptime(dob1,'%Y-%m-%d %H:%M:%S')
    user['dob']=datetime.datetime(dob.year,dob.month,dob.day).strftime("%d/%m/%Y")
    roles = []
    role1 = {}
    cityTenantId= sheet.cell(rowIndex, 2).value
    cityTenantId=cityTenantId.replace(' ','')
    cityTenantId='pb.'+cityTenantId.lower()
    multiRole= sheet.cell(rowIndex, 10).value
    multiRoleList=multiRole.split(",")
    #print(multiRoleList)
    if len(multiRoleList) > 1:
        for roleCode in multiRoleList :
            roleTemp = {}
            roleTemp['code']=roleCode
            roleTemp['tenantId'] = cityTenantId
            #print("in loop"+roleTemp['code'])
            roles.append(roleTemp)
    else :
        role1['code']=multiRole
        role1['tenantId'] = cityTenantId
        roles.append(role1)
        #print("from else" + role1['code'])
    role2 = {}
    if sheet.cell(rowIndex, 12).value is not None:
        role2['tenantId'] = cityTenantId
        role2['code'] = sheet.cell(rowIndex, 12).value
        roles.append(role2)
    user['roles'] = roles
    user['password'] = 'mGram@123'
    #'''.join(random.sample(string.ascii_lowercase + string.digits + string.ascii_uppercase + string.punctuation , k=8))
    #print(user['password']);
    user['type'] = 'EMPLOYEE'
    user['active'] = True
    user['status'] = 'ACTIVE'
    user['permanentAddress']='test'
    user['correspondenceAddress']='test'
    user['otpReference']='test'
    user['userName']=sheet.cell(rowIndex, 4).value
    requestInfo = {'authToken': accesstoken}

    userRequest = {}
    userRequest['requestInfo']=requestInfo
    userRequest['user']=user

    userSearchResponse=userExist(user['mobileNumber'],user['name'],requestInfo)
    #print("Search Result :\n" )
    #print(userSearchResponse);

    if (userSearchResponse==None or  len(userSearchResponse) == 0):
          post_response = requests.post(url=host
                                      + 'user/users/_createnovalidate',
                                      headers={'Content-type': 'application/json'},
                                      json=userRequest)
          jsondata = post_response.json()
          #print("creating user result");
          #print(jsondata);
          #print("\n");
          # DuplicateUserNameException is thrown if the mobilenumber exists but the name is different
          if( jsondata.get('user') != None ):
                print("user created , adding sms for "+user['name'] +" with id " )
                print(jsondata.get('user')[0]['id'])
                sheet.cell(rowIndex, 13).value=jsondata.get('user')[0]['id']
                # addSms(user['name'],user['mobileNumber'], user['password'], loginUrl, iphindi, file)
                addSms(user['name'],user['mobileNumber'], user['password'], loginUrl, ippunjabi, file)
                #addSms(user['name'], user['mobileNumber'], user['password'], loginUrl, ipenglish, file)
    else:
          print("updating user...")
          for role in  userSearchResponse[0]['roles']:
              del role['name']

          for role in user['roles']:
            if role in userSearchResponse[0]['roles']:
                print(role['code'] +  ' Role already exist for tenantid '+role['tenantId'])
            else:
               #print(role['code'] + ' adding role for tenantid '+role['tenantId'])
                userSearchResponse[0]['roles'].append(role)
        # print(userSearchResponse[0])

          userRequest['user']=userSearchResponse[0]
          dob3 = str(sheet.cell(rowIndex, 8).value)
          dob2 = datetime.datetime.strptime(dob3, '%Y-%m-%d %H:%M:%S')
          userRequest['user']['dob']=datetime.datetime(dob2.year,dob2.month,dob2.day).strftime("%d/%m/%Y")
          post_response = requests.post(url=host
                                          + 'user/users/_updatenovalidate',
                                          headers={'Content-type': 'application/json'},
                                          json=userRequest)
          jsondata = post_response.json()
          print(jsondata);
          print("updated "+str(user['userName']));


def pushSms():
    result = subprocess.run(cmd,shell=True,stdin=subprocess.PIPE)
    print('done');
def addSms(userName,phno,pwd,url,template,file):
    ipTemp = template
    ipTemp = ipTemp.replace("{USER}", userName)
    ipTemp = ipTemp.replace("{PASSWORD}", pwd)
    ipTemp = ipTemp.replace("{PHNO}", str(phno))
    ipTemp = ipTemp.replace("{USERNAME}", str(phno))
    ipTemp = ipTemp.replace("{LINK}", url)
    print(ipTemp);
    file.write(ipTemp)
    file.write("\n")

def userExist(mobilenumber,name,requestInfo):
     userSearchRequest={}
     userSearchRequest['RequestInfo']=requestInfo
     userSearchRequest['name']=name
     userSearchRequest['mobileNumber']=mobilenumber
     userSearchRequest['tenantId']=stateTenantId
     userSearchRequest['userType']='EMPLOYEE'
     post_response = requests.post(url=host
                                  + 'user/_search',
                                  headers={'Content-type': 'application/json'},
                                  json=userSearchRequest)
     jsondata = post_response.json()
     print("search user result");
     print(jsondata);
     return jsondata.get('user')



def accessToken():
    query = {
        'username': username,
        'password': password,
        'userType': 'EMPLOYEE',
        'scope': 'read',
        'grant_type': 'password',
        }
    query['tenantId'] = 'pb'
    response = requests.post(host + 'user/oauth/token', data=query,
                             headers={
        'Connection': 'keep-alive',
        'content-type': 'application/x-www-form-urlencoded',
        'origin': host,
        'Authorization': 'Basic ZWdvdi11c2VyLWNsaWVudDo=',
        })
    jsondata = response.json()
    print(jsondata)
    return jsondata.get('access_token')


if __name__ == '__main__':
    connect()
